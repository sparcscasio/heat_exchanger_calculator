#
import math
import matplotlib
matplotlib.use("Agg")  # 화면 표시 없이 이미지 생성 (macOS 안전)
import matplotlib.pyplot as plt
import matplotlib.pyplot as plt
import numpy as np
import sys
import time
import random
import pandas as pd
import tkinter
import tkinter.font
import openpyxl
from datetime import datetime
from flask import Flask, request, jsonify
import threading
import webview
from flask_cors import CORS
import io, base64
# from numba import jit

app = Flask(__name__)
CORS(app)

@app.route("/api/calc", methods=["POST"])
def calc():
    data = request.get_json()
    arr = np.array([1, 2, 3])
    squared = np.square(arr)
    df = pd.DataFrame({"x": arr, "y": squared})
    df.to_excel("out.xlsx", index=False)
    return jsonify({"result": squared.tolist()})

@app.route("/api/simulate", methods=["POST"])
def simulate():
   inputs = request.get_json()
   res = {}
   ws = {}

   # 입력값을 변수에 할당한다.
   Cold_Q = inputs.get('Cold_Q')
   La_Heat = inputs.get('La_Heat')
   tube_ea = inputs.get('tube_ea')
   tube_L = inputs.get('tube_L')
   Hot_Q = inputs.get('Hot_Q')
   d = inputs.get('d')   # 판 간격 .004 m 4mm
   Hot_intemp1 = inputs.get('Hot_intemp1')
   Cold_Spec = inputs.get('Cold_Spec')
   Temp_in = inputs.get('Temp_in')
   N = inputs.get('N')
   Tk = inputs.get('Tk')
   Fouling_F = inputs.get('Fouling_F')
   A = inputs.get('A')
   B = inputs.get('B')
   t_pitch = inputs.get('t_pitch')
   C = inputs.get('C')
   D = inputs.get('D')
   E = inputs.get('E')
   L = inputs.get('L')
   M = inputs.get('M')
   Hot_T = inputs.get('Hot_T')
   Temp_low = inputs.get('Temp_low')
   Hot_spec = inputs.get('Hot_spec')

   # 온도입력
   H_1 = inputs.get('H_1')
   H_2 = inputs.get('H_2')
   H_3 = inputs.get('H_3')
   H_4 = inputs.get('H_4')
   H_5 = inputs.get('H_5')
   H_6 = inputs.get('H_6')
   H_7 = inputs.get('H_7')

   # VO Property 
   VO1 = inputs.get('V01')
   VO2 = inputs.get('V02')
   VO3 = inputs.get('V03')
   VO4 = inputs.get('V04')
   VO5 = inputs.get('V05')
   VO6 = inputs.get('V06')
   VO7 = inputs.get('V07')
   VO_IN = VO1

   # CP Property 
   C_CP1 = inputs.get('C_CP1')
   CP2 = inputs.get('CP2')
   CP3 = inputs.get('CP3')
   CP4 = inputs.get('CP4')
   CP5 = inputs.get('CP5')
   CP6 = inputs.get('CP6')
   CP7 = inputs.get('CP7')
   CP_IN = C_CP1

   # TC Property 
   C_TC1 = inputs.get('C_TC1')
   TC2 = inputs.get('TC2')
   TC3 = inputs.get('TC3')
   TC4 = inputs.get('TC4')
   TC5 = inputs.get('TC5')
   TC6 = inputs.get('TC6')
   TC7 = inputs.get('TC7')
   TC_IN = C_TC1

   # PR Property 
   C_PR1 = inputs.get('C_PR1')
   PR2 = inputs.get('PR2')
   PR3 = inputs.get('PR3')
   PR4 = inputs.get('PR4')
   PR5 = inputs.get('PR5')
   PR6 = inputs.get('PR6')
   PR7 = inputs.get('PR7')
   PR_IN = C_PR1

   # VI Property 
   C_VI1 = inputs.get('C_VI1')
   VI2 = inputs.get('VI2')
   VI3 = inputs.get('VI3')
   VI4 = inputs.get('VI4')
   VI5 = inputs.get('VI5')
   VI6 = inputs.get('VI6')
   VI7 = inputs.get('VI7')
   VI_IN = C_VI1

   # 계산
   Req = Cold_Q * La_Heat
   Q = Req
   Tu_t = Tk*2   #튜브 두께 

   QWCP = .81
   TCL = .86
   DET = .0002  # 수렴

   Hot = Hot_Q/3600.
   Hot_V = Hot//0.029/943
   Hot_V = Hot//0.023/943
   H_TC = 0.35*1.16   # w/m-k
   H_TC = 0.35   # kcal

   TOC = 0  # Total_Convec
   TCP = 0  # Specific heat

   T1 = Temp_in
   T2 = Hot_intemp1  
   pai = np.pi

   # Water Liquid COLD Liquid
   C_VO = -1.38931902e-13* T1**4 + 3.39458511e-11 *T1**3 + 7.23194774e-10 * T1**2 + 1.87586756e-07* T1 + 9.97692559e-04
   C_CP = 1.0
   C_TC =(-1.17187500e-07* T1**4 + 5.14467593e-05*T1**3 +-1.49288194e-02*T1**2 + 2.28990741e+00*T1 + 5.57883333e+02)/1000*TCL
   C_VI = (7.81250000e-11* T1**4 +-3.07638889e-08*T1**3 + 4.81041667e-06*T1**2 +-3.81575397e-04*T1 + 1.59400000e-02)/10000
   C_PR =  7.91666667e-08* T1**4 +-2.98101852e-05*T1**3 + 4.39784722e-03*T1**2 +-3.22542791e-01*T1 + 1.19163333e+01

   # Glycol Water / 5 barG / 20~80
   H_VO=1/(-9.72222222225747E-10*T1**6 +2.95833333334354E-07*T1**5 +-0.000036180555555674*T1**4 +0.00226875000000702*T1**3 +-0.0788722222224456*T1**2 +0.985166666670283*T1 +1064.59999999998)    #VO 단위질량당부피 [m3/kg]
   H_CP=(9.72222222237005E-14*T1**6 +-2.87500000004451E-11*T1**5 +3.28472222227577E-09*T1**4 +-1.53958333336613E-07*T1**3 +-2.02944444433701E-06*T1**2 +0.00117368333333155*T1 +0.751150000000012)    #CP 비열 [kcal/kgK]
   H_TC=(-8.35960638194554E-13*T1**6 +2.4720550300897E-10*T1**5 +-2.93183338109684E-08*T1**4 +1.77522212668052E-06*T1**3 +-6.02584312600458E-05*T1**2 +0.00169876755517174*T1 +0.315219260533114)   #Thermal Conduct 열전도도 [kcal/mK]
   GW_cP=(-1.90277777778148E-11*T1**6 +4.38750000001133E-09*T1**5 +-2.82152777779164E-07*T1**4 +-1.03604166665801E-05*T1**3 +0.00251590555555266*T1**2 +-0.166062833333284*T1 +5.55809999999967)   #Viscosity 점도 [cP]
   H_VI=(-9.44444444448553E-18*T1**6 +1.7166666666791E-15*T1**5 +4.72222222071443E-15*T1**4 +-2.45249999999067E-11*T1**3 +2.73962222221913E-09*T1**2 +-1.57384666666615E-07*T1 +5.15039999999966E-06)   #Kinetic Viscosity 동점성계수 [m2/s]
   H_PR=(-5.63150143643295E-11*T1**6 +6.98315976359687E-09*T1**5 +1.06381819268007E-06*T1**4 +-0.000287155332857528*T1**3 +0.0273229496066831*T1**2 +-1.48490105059018*T1 +46.6024012145047)   #Prantl number 

   # Boiling Stage
   res['SP'] = round(C_CP,3)
   res['PR'] = round(C_PR,3)
   res['VI'] = round(C_VI,9)
   res['TC'] = round(C_TC,3)
   res['VO'] = round(C_VO,3)

   C_CP1 = C_CP 
   C_PR1 = C_PR
   C_VI1 = C_VI
   C_TC1 = C_TC
   VO1 = C_VO
   SG1 = 1/C_VO/1000

   H_CP1 = H_CP
   H_TC1 = H_TC
   H_VO1 = H_VO
   GW_cP1 = GW_cP
   GW_PR1 = H_PR
   H_VI1 = H_VI
   Hd = t_pitch

   Surface_area = pai * tube_L**2 / 4  * tube_ea

   Cold_V = Cold_Q * C_VO / 3600 / (tube_ea * pai * (d-Tu_t)**2/ 4)
   Volume_Rate1 = Cold_Q*C_VO

   Cold_intemp = Temp_in
   Outlet_Temp = Cold_intemp

   Hot_intemp_LMTD = Hot_intemp1
   Cold_Re = Cold_V * d / C_VI
   Nu = 0.023 * Cold_Re**.8 * C_PR**0.4

   Cold_h = Nu * C_TC / d   
   Hot_Red = Hot_V * d / 0.0000016
   Hot_Nu = 0.664 * Hot_Red**.5 * H_PR**.3333
   Hot_h = Hot_Nu * H_TC / t_pitch 

   Total_convec = 1/(1/Hot_h + 1/Cold_h + 0.0005 )
   res['Total_convec'] = Total_convec
   Total_convec = 110

   surface_a = Q / (Total_convec*(Hot_intemp1-Cold_intemp))
   surface_a = Q / (Total_convec)
   surface_a1= surface_a

   Hot_outtemp =  Hot_intemp1 - Q/(Hot_Q*H_CP) 
   BGW_outtemp = Hot_outtemp 
   TQ1 = Q

   res['Boiling surface Area m^2'] = round(surface_a,2)
   res['Heat Capa.(Kcal/h)'] = round(Q/1000,2)
   res['Cold Inlet Temp'] = Cold_intemp
   res['Cold Outlet Temp'] = Cold_intemp
   res['GW Inlet Temp'] = Hot_intemp1
   res['GW Outlet Temp'] = round(Hot_outtemp,2)

   GW_intemp = Q/(Hot_Q*H_CP)

   T_surface = 0
   TOQ = Q
   Hot_intemp = Hot_outtemp

   N = int(N)
   BLNG_intemp = Temp_in

   DPT = 0
   DpgT = 0
   DP_in = 0
   DPL = DP_in 

   Del_T = (Hot_intemp - Cold_intemp)

   R = 1/Hot_h + 1/Cold_h 
   SKT1 = Hot_intemp - Del_T * (1/Hot_h)/R
   SKT_out = Hot_intemp - Del_T * (1/Hot_h)/R
   Minim = 30
   Hd = t_pitch
   Total_convec1 = Total_convec
   SKT_out = Temp_in
   CPP = C_CP
   C_CP1 = C_CP 

   fig = plt.figure(figsize=(13,8))
   plt.axis ([Surface_area*-.15, Surface_area*1.15, Temp_low, Hot_T+20])

   start_T = time.time()
   Int_T = time.time()

   # prevent unassociated error
   Total_convec3 = 0
   ToL_Spec = 0
   Int_Hot_intemp = 0
   Int_Cold_intemp = 0
   T_surface2 = 1
   Int_Cold_intemp2 = 0

   for k in range (0, 20, 1) :   # GW 출구 Temp  예상
      TT2 =  Hot_intemp1 + k/10
      T_surface = 0
      c1 = 0
      Total_convec = 650
      DPT = 0
      DpgT = 0
      Hot_intemp = TT2 
      Cold_intemp = Temp_in 
      TOQ = 0
      TOC = 0
      SKT_out = Temp_in

      for i  in range (N, 360 - 2 * N, 1) :
         ii = i/2
         c1 = Total_convec 
         CPP1 = C_CP  
         Total_convec_0 = Total_convec            
         surface_a = 2*tube_ea* math.sin(math.radians(ii))*tube_L/2 * (math.cos(math.radians(ii))*tube_L/2- math.cos(math.radians(ii+1/2))*tube_L/2)      
         T_surface =  T_surface + surface_a     

      # Water Liquid COLD Liquid
         T1 = Cold_intemp    
         C_VO = -1.38931902e-13* T1**4 + 3.39458511e-11 *T1**3 + 7.23194774e-10 * T1**2 + 1.87586756e-07* T1 + 9.97692559e-04
         C_CP = 1.0
         C_TC =(-1.17187500e-07* T1**4 + 5.14467593e-05*T1**3 +-1.49288194e-02*T1**2 + 2.28990741e+00*T1 + 5.57883333e+02)/1000*TCL
         C_VI = (7.81250000e-11* T1**4 +-3.07638889e-08*T1**3 + 4.81041667e-06*T1**2 +-3.81575397e-04*T1 + 1.59400000e-02)/10000
         C_PR =  7.91666667e-08* T1**4 +-2.98101852e-05*T1**3 + 4.39784722e-03*T1**2 +-3.22542791e-01*T1 + 1.19163333e+01
         C_cP = C_VI * 1000 / C_VO      # Centi-Poise

         T2 = Hot_intemp       
         x =  [H_1,     H_2,    H_3,    H_4,    H_5,    H_6,    H_7] 
         y =  [VO_IN,   VO2,    VO3,    VO4,    VO5,    VO6,    VO7 ] # 비체적 VO 3bar LN2

         def H_VO(x, y, T2) :  
            poly = np.polyfit(x, y, 6)
            H_VO = 1/(math.pow(T2,6)*poly[0]+math.pow(T2,5)*poly[1]+math.pow(T2,4)*poly[2]+math.pow(T2,3)*poly[3]+math.pow(T2,2)*poly[4]+math.pow(T2,1)*poly[5]+poly[6]) 
            return H_VO
         H_VO = H_VO(x, y, T2)

         y = [CP_IN ,CP2, CP3, CP4, CP5, CP6, CP7] # Cp 
         def H_CP(x, y, T2) :  
            poly = np.polyfit(x, y, 6)
            H_CP = (math.pow(T2,6)*poly[0]+math.pow(T2,5)*poly[1]+math.pow(T2,4)*poly[2]+math.pow(T2,3)*poly[3]+math.pow(T2,2)*poly[4]+math.pow(T2,1)*poly[5]+poly[6])    
            return H_CP
         H_CP = H_CP(x,y,T2)

         y = [TC_IN, TC2, TC3, TC4, TC5, TC6, TC7] # TC 
         def H_TC(x, y, T2) :  
            poly = np.polyfit(x, y, 6)   
            H_TC = (math.pow(T2,6)*poly[0]+math.pow(T2,5)*poly[1]+math.pow(T2,4)*poly[2]+math.pow(T2,3)*poly[3]+math.pow(T2,2)*poly[4]+ math.pow(T2,1)*poly[5]+poly[6])*.86
            return H_TC
         H_TC = H_TC(x, y, T2) 

         y = [PR_IN, PR2, PR3, PR4, PR5, PR6, PR7] # PR 
         def H_PR(x, y, T2) : 
            poly = np.polyfit(x, y, 6)      
            H_PR =(math.pow(T2,6)*poly[0]+math.pow(T2,5)*poly[1]+math.pow(T2,4)*poly[2]+math.pow(T2,3)*poly[3]+math.pow(T2,2)*poly[4]+math.pow(T2,1)*poly[5]+poly[6])
            return H_PR
         H_PR = H_PR(x, y, T2)        

         y = [VI_IN, VI2, VI3, VI4, VI5, VI6, VI7] # VI 
         def H_VI(x, y, T2) :
            poly = np.polyfit(x, y, 6)
            H_VI = (math.pow(T2,6)*poly[0]+math.pow(T2,5)*poly[1]+math.pow(T2,4)*poly[2]+math.pow(T2,3)*poly[3]+math.pow(T2,2)*poly[4]+math.pow(T2,1)*poly[5]+poly[6])/10000
            return H_VI
         H_VI = H_VI(x,y,T2) 

         H_cP = H_VI * 1000 / H_VO # Centi-Poise       
         Cold_V1 = Cold_Q / 3600           # 물
         Hd = t_pitch
         Area = tube_ea / 2 / L * Hd * math.sin(math.radians(ii)) * tube_L 

         Cold_V = Cold_V1/(Area * 1/C_VO)
         Cold_Re = Cold_V * Hd / C_VI   

         Nu = 0.023 * math.pow(Cold_Re,.8) * math.pow(C_PR,0.333)
         if Cold_Re <  50000  :                                               # Tube 실 직경
            Nu = 0.664 * math.pow(Cold_Re,.5) * math.pow(C_PR,0.333)  
         
         Cold_h = Nu * C_TC / Hd                         
         Hot = Hot_Q  / 3600                                                      # 메탄
         Area = tube_ea / 2 / M * Hd * math.sin(math.radians(ii))  * tube_L       # GW 유로면적

         Hot_V = Hot / (Area * 1/H_VO)  
         Hot_Red = Hot_V * Hd / H_VI 

         Hot_Nu = 0.023 * math.pow(Hot_Red,.79) * math.pow(H_PR,.33)
         if  Hot_Red <  50000 : 
            Hot_Nu = 0.664 * math.pow(Hot_Red,.5) * math.pow(H_PR,.33)
         Hot_h = Hot_Nu * H_TC / Hd 

         if i == 158 : 
            Cold_VC =   Cold_V 
            Hot_VC = Hot_V
            C_cPC =  C_cP
            H_cPC = H_cP

         if i == 316 : 
            Cold_VCT =   Cold_V 
            Hot_VCT = Hot_V
            C_cPT =  C_cP
            H_cPT = H_cP

         Total_convec = 1/(1/Cold_h + 1/Hot_h + Fouling_F + Tk/14.4)        # 평행 평판 

         if i <= N :
            Total_convec_0 = 1/(1/Cold_h + 1/Hot_h + Fouling_F + Tk/14.4) * .85
            R = 1/Hot_h + 1/Cold_h + Fouling_F + Tk/14.4

         DP_in = 0

         for j in range(100000) :
            j = (j + 1)/150  
            TQ = (Total_convec)*surface_a*j

            Hot_outtemp =  Hot_intemp + TQ / (Hot_Q * H_CP)   # Counter Flow
            Cold_outtemp = Cold_intemp + TQ / (Cold_Q * C_CP) 
            
            hdt = Hot_intemp - Cold_outtemp           
            cdt = Hot_outtemp - Cold_intemp  

            LMTD = (hdt-cdt)/math.log(hdt/cdt)        
            TTQ = Total_convec * surface_a * LMTD
            DT = math.pow((TTQ - TQ), 2) /1000

            if DT <= DET :
               Volume_Rate = Cold_Q * C_VO
               TOQ = TOQ + TQ
               ToL_Q = TOQ
               ToL_W1 = ToL_Q  
               Dpg = (64/Cold_Re * (tube_L + 100*Hd)                   / N) * (1/C_VO) * (Cold_V*L)**2 /(2 * Hd) # glycol side(수력학적직경)
               Dp = ( 64/Hot_Red* (tube_L + 100*Hd)                  / N) * (1/H_VO) * (Hot_V*L)**2 / (2 * Hd) # glycol side(수력학적직경)
               DPT = DPT + Dp      # 유체 속도 및 비중량이 길이 방향에 따라 변하기 때문에 길이방향을 N개(60~120)로 미분하여 계산하고 적분함.
               DpgT = DpgT + Dpg
               DP_in = (DPT - Dp) / 10000
               DP_out = DPT / 10000
                     
               Del_T = (Hot_outtemp - Cold_outtemp)
               SKT = Cold_outtemp + Del_T * Hot_h/(Cold_h + Hot_h)
               SKT_in = SKT_out           
               SKT_out =  SKT   

               TOC = TOC + Total_convec
               TCP = TCP + C_CP   

               Location = 0                         # 비열
               DTT = Location                       # Start Scale 변경, 비열입구 온도 위치 

               CPP1 = CPP1*100 - C_CP1*100 + Location
               CPP2 = C_CP*100 - C_CP1*100 + Location

               Lc = 140                             #  열전달계수 입구위치 
               Scailc = Total_convec1 - Lc

               c1 = c1/6  - Scailc
               c2 = Total_convec/6  - Scailc       
                     
               x = np.arange (0,400,1)
               x = T_surface 
               y = Cold_intemp
               y1 = Cold_outtemp 
               z = Hot_intemp
               z1 = Hot_outtemp

               a = DP_in * 20
               a1 = DP_out * 20
               s1 = SKT_in
               s2 = SKT_out

               line_Tk = .3             
               plt.plot ([T_surface-surface_a, T_surface ],[y, y1 ] , c = 'b', lw = line_Tk+.4) 
               plt.plot ([T_surface-surface_a, T_surface ],[z, z1 ] , c = 'r', lw = line_Tk+.5) 
               if k == 6 : 
                  plt.plot ([T_surface-surface_a, T_surface ],[z,  z1 ], c = 'k', lw = line_Tk+.6) 

               plt.plot ([T_surface-surface_a, T_surface ],[a, a1 ],  c = 'firebrick', lw = line_Tk)    # DP_in Temp 

               plt.plot ([T_surface-surface_a, T_surface ],[s1, s2 ], c = 'g', lw = line_Tk)    # SKin Temp
               plt.plot ([T_surface-surface_a, T_surface ],[c1, c2 ], c = 'm', lw = line_Tk)    # Convection
               plt.plot ([T_surface-surface_a, T_surface ],[CPP1, CPP2 ], c = 'firebrick', lw = .5)   # Specific heat 

               Cold_intemp = Cold_outtemp
               Hot_intemp = Hot_outtemp 
               DP_in = DP_out                             

               if math.pow ((Hot_spec - Hot_outtemp),2) < .0005 :  # Hot 출구 예상온도 46도
                  Int_Cold_intemp =  Cold_intemp  
                  Int_Hot_intemp = Hot_outtemp
                  T_surface2 = T_surface 
                  Total_convec3 = Total_convec 
                  DPOUT1 =  DP_out

               if math.pow((Cold_Spec - Cold_intemp),2) < .0001 :     # LNG 출구 예상온도 42도  
                  Int_Cold_intemp2 =  Cold_intemp  
                  Int_Hot_intemp2 = Hot_intemp     
                  T_surface2 = T_surface
                  Margine = (Surface_area - T_surface2)*100/ Surface_area
                  Margine = (Surface_area - T_surface)/ T_surface
                  ToL_W1 = ToL_Q / 859.8
                  DPOUT2 =  DP_out 
                  ToL_Spec = ToL_W1  
                  DPOUT2 =  DP_out
                  break
                     
               if k==1 :
                  Q1 = ToL_Q/859.8
                  TG1 = Hot_outtemp  
               if k==2 : 
                  Q2 = ToL_Q/859.8 
                  TG2 = Hot_outtemp         
               if k==3 :
                  Q3 = ToL_Q/859.8 
                  TG3 = Hot_outtemp  
               if k==4 :
                  Q4 = ToL_Q/859.8  
                  TG4 = Hot_outtemp   
               if k==5 :
                  Q5 = ToL_Q/859.8  
                  TG5 = Hot_outtemp     
               if k==6 :
                  Q6 = ToL_Q/859.8  
                  TG6 = Hot_outtemp    
               if k==7 :
                  Q7 = ToL_Q/859.8  
                  TG7 = Hot_outtemp      
               Int_T = time.time()
               break   
      if (Hot_outtemp > Hot_T ) : break

   # ==== 그래프 그리는 부분 ==== #
   plt.rc('font', size=12)   
   plt.text(T_surface*.996, Hot_outtemp - 22, round(TG1,1), c = 'r')        # 100% 출구온도
   TOC = TOC / (320 - 2 * N)
   TCP = TCP / (320 - 2 * N) 

   plt.plot ([0],[surface_a1] , c = 'b',           label='COLD'         )            
   plt.plot ([0],[surface_a1] , c = 'r',           label='HOT')
   plt.plot ([0],[surface_a1] , c = 'firebrick',   label='P. Loss'  )
   plt.plot ([0],[surface_a1] , c = 'm',           label='Kcal/m2hr'   ) 

   plt.rc('font', size=12)
   plt.legend(loc=(.3,.02), ncol=1, frameon=True, shadow=True)
   plt.tick_params(axis='both', direction='in', length=6, pad=7, labelsize=8)
   plt.scatter (T_surface*1.05, Hot_intemp_LMTD*1.1 , c = 'r', alpha = .01)  # 그리드 크기 조정

   plt.rc('font', size=20)
   plt.xlabel('Surface Area (m^2)',fontsize=13)
   plt.ylabel('Temperature C',     fontsize=13) 

   plt.title(A,fontsize=13) 
   plt.rc('font', size=12) 
   plt.text(T_surface*-.1,Temp_in-10,'Cold_In', c = 'b')
   plt.text(T_surface*-.1,Temp_in-5,  round(Temp_in,1), c = 'b'  )                           # Cold 입구온도

   Sd = 6

   DTS = Total_convec1 - Lc 
   Total_convec2 = Total_convec_0

   plt.rc('font', size=12) 
   plt.text(T_surface*-0.1, (Total_convec2 - DTS)/8 + 30,  'Kcal/m2h',              c = 'm')   # Heat transfer Coef.  - -
   plt.text(T_surface*-0.1, (Total_convec2 - DTS)/8 + 36,   round(Total_convec2),   c = 'm') 
   plt.text(T_surface*1.02, (Total_convec  - DTS)/8 + 30,  'Kcal/m2h',              c = 'm')
   plt.text(T_surface*1.02, (Total_convec -  DTS)/8 + 36,   round(Total_convec),    c = 'm') 
   plt.text(T_surface*.6,   (Total_convec3 - DTS)/8 + 40,   round(Total_convec3),   c = 'm')
   plt.text(T_surface*-.1,Hot_intemp_LMTD-1, round(Hot_intemp_LMTD,1), c = 'r' )  # GW 입구온도
   plt.text(T_surface*.002,  TT2+2.5, round(TT2,1), c = 'r' )  # GW 입구온도
   plt.text(T_surface*-.1, -5, ' Press Loss', c = 'k' )                           # Press Loss
   plt.text(T_surface*-0.1,Hot_intemp_LMTD+5,'Hot_out', c = 'r')

   Hot_intemp_LMTD = Hot_intemp  # Pararell
   plt.rc('font', size=14)
   aa = 1.2
   COR = 50
   plt.text(T_surface*.2,(Hot_intemp_LMTD-Temp_in)/aa + COR, 'Cold ')     
   plt.text(T_surface*.3,(Hot_intemp_LMTD-Temp_in)/aa + COR, round(Cold_Q), c = 'b') 
   plt.text(T_surface*.4,(Hot_intemp_LMTD-Temp_in)/aa + COR, 'Kg/h')
   bb = aa*1.15
   plt.text(T_surface*.2,(Hot_intemp_LMTD-Temp_in)/bb+ COR, 'HOT ') 
   plt.text(T_surface*.3,(Hot_intemp_LMTD-Temp_in)/bb+ COR, round(Hot_Q), c = 'r') 
   plt.text(T_surface*.4,(Hot_intemp_LMTD-Temp_in)/bb+ COR, 'Kg/h ')
   cc = bb*1.18
   plt.text(T_surface*.2,(Hot_intemp_LMTD-Temp_in)/cc+ COR, 'Spec ') 
   plt.text(T_surface*.3,(Hot_intemp_LMTD-Temp_in)/cc+ COR, round(ToL_Spec),   c = 'k')
   plt.text(T_surface*.4,(Hot_intemp_LMTD-Temp_in)/cc+ COR, 'Kw ')

   plt.text(T_surface*.5,(Hot_intemp_LMTD-Temp_in)/cc+ COR, 'Toq :')
   plt.text(T_surface*.57,(Hot_intemp_LMTD-Temp_in)/cc+ COR, round(ToL_Q/860),  c = 'k')

   dd = cc*1.21
   plt.text(T_surface*.2,(Hot_intemp_LMTD-Temp_in)/dd+ COR, 'C.dp')     
   plt.text(T_surface*.3,(Hot_intemp_LMTD-Temp_in)/dd+ COR, round(DpgT/1000,1), c = 'b') 
   plt.text(T_surface*.4,(Hot_intemp_LMTD-Temp_in)/dd+ COR, 'Kpa')
   ee = dd*1.28
   plt.text(T_surface*.2,(Hot_intemp_LMTD-Temp_in)/ee+ COR, 'H.dp')     
   plt.text(T_surface*.3,(Hot_intemp_LMTD-Temp_in)/ee+ COR, round(DPT/1000,1), c = 'firebrick')  
   plt.text(T_surface*.4,(Hot_intemp_LMTD-Temp_in)/ee+ COR, 'Kpa')
   ff = ee*1.4
   plt.text(T_surface*.2, (Hot_intemp_LMTD-Temp_in)/ff+ COR, 'Fouling F')  
   plt.text(T_surface*.35,(Hot_intemp_LMTD-Temp_in)/ff+ COR, round(Fouling_F,7), c = 'm') 

   plt.rc('font', size=12)
   plt.text(Surface_area*.2, Int_Hot_intemp+3,  round(Int_Hot_intemp,1), c = 'r')  # 80% 면적온도
   plt.text(Surface_area*.2, Int_Cold_intemp-5, round(Int_Cold_intemp,1), c = 'b')

   plt.text(T_surface2,  Int_Cold_intemp2-8 , round(Int_Cold_intemp2,1), c = 'b')

   plt.text(T_surface*1.02, Hot_outtemp + 0,  round(Hot_outtemp,1), c = 'r')        # 100% 출구온도
   plt.text(T_surface*1.02, Hot_outtemp + 6,  'Hot_In' , c = 'r')
   plt.text(T_surface*1.02, Cold_outtemp - 2, round(Cold_outtemp,1), c = 'b')
   plt.text(T_surface*1.02, Cold_outtemp + 4, 'Cold_Out' , c = 'b')
   plt.text(T_surface*1.02, SKT + 4,          round(SKT,1) , c = 'g')
   plt.text(T_surface*1.02, SKT + 10,         'Skin_Temp' , c = 'g')
   plt.text(T_surface*1.02, DP_out*10, round(DPT/1000,1), c = 'firebrick')
   plt.text(T_surface*1.07, DP_out*10, ' KPa',           c = 'firebrick')

   plt.rc('font', size=12)

   surface_pe = round(T_surface2*100/T_surface,1)                                  # 출구 온도 Spec.

   plt.gca().set_facecolor('oldlace')
   plt.rc ('font', family = "consolas", weight='bold', size=10)
   plt.text(T_surface*.88, Temp_low+10,'Coded by KIM & Mytec')

   plt.grid(True)
   buf = io.BytesIO()
   fig.savefig(buf, format="png")
   buf.seek(0)
   plt.close(fig)
   # Base64 인코딩
   img_base64 = base64.b64encode(buf.read()).decode("utf-8")

   ## ==== 여기까지 그래프 그리기 ==== ##

   ## ==== 여기는 결과 출력 ==== ##
   ToTal_K = TOC  # Average Heat Transfer Coef.
   ToL_Q = TOQ
   ToL_W = ToL_Q*1.163
   ToTal_K_W = ToTal_K * 1.163 

   margine = (100 - surface_pe) / surface_pe * 100
   mardumi = 1/(1 + margine/100)
   service = ToTal_K_W * mardumi
   LMTD = ToL_Spec *1000 / service / T_surface  
   LMTD = ToL_W / service / (T_surface * .99)

   res['Heat Capacity (kcal/h)'] = f"{round(ToL_Q/1000,1)} ({round(ToL_Q/859.8,1)} KW)"
   res['Surf. Area ( m^2 )'] = round(T_surface,2)
   res['Heat Transfer Coef. (Kcal/m2.hr)'] = f"{round(ToTal_K,1)} ({round(ToTal_K_W,1)} W/m2.K)"
   res['Transfer Service (W/m2.K)'] = round(service,1)
   res['Spec. Heat Capa (KW)'] = round(ToL_Spec,1)
   res['Fouling Factor'] = round(Fouling_F,5)

   res['ToTal LMTD'] = round(LMTD,2)
   res['Cold In Temp'] = round(Temp_in,1)
   res['Cold out Temp'] = round(Cold_outtemp,1)
   res['HOT IN Temp'] = round(Hot_outtemp,1)
   res['HOT out Temp'] = round(Hot_intemp1,1)
   res['Hot_V (m/s)'] = round(Hot_VC,1)
   res['Press Loss (Kpa)'] = round(DPT/1000,2)

   res['Sface_margin (%)'] = round(margine, 1)
   res['Cold In_V. (m/s)'] = round(Cold_V,1)
   res['Cold Out_V (m/s)'] = round(Cold_VC,1)
   res['Cold_Q (kg/h)'] = round(Cold_Q)
   res['HOT_Q (kg/h)'] = round(Hot_Q)
   res['Tube_ea (ea)'] = round(tube_ea)
   res['Tube Dia. (mm)'] = round(d*1000,2)

   res['Spec.H'] = f'In: {round(C_CP1,2)} PR: {round(C_PR1,2)} VI: {round(C_VI1,7)} TC: {round(C_TC1,3)} VO: {round(VO1,3)}'
   res['Cold'] = f'Out: {round(C_CP,3)} PR: {round(C_PR,2)} VI: {round(C_VI,8)} TC: {round(C_TC,3)} VO: {round(C_VO,3)}'

   res['HOT_CP'] = f'In: {round(H_CP1,3)} HOT_VI: {round(H_VI1,8)} HOT_TC: {round(H_TC1,3)} HOT_VO : {round(H_VO1,3)}'
   res['HOT'] = f'Out: {round(H_CP,3)} VI: {round(H_VI,8)} TC: {round(H_TC,3)} VO: {round(H_VO,3)}'
   res['Vol_Rate in(m3/h)'] = round(Volume_Rate1,1)
   res['Vol_Rate out(m3/h)'] = round(Volume_Rate,1)

   res['Heat Capacity (KW)'] = f'{round(Q1,1)}, {round(Q2,1)}, {round(Q3,1)}, {round(Q4,1)}, {round(Q5,1)}'
   res['Hot In'] = f'{round(TG1,1)}, {round(TG2,1)}, {round(TG3,1)}, {round(TG4,1)}, {round(TG5,1)}'

   ws['Tube_ea'] = tube_ea
   ws['Tube_L'] = tube_L
   ws['Tube Dai'] = d
   ws['Tube Thickness'] = Tk
   ws['Gw_In Temp'] = Hot_intemp_LMTD
   ws['Gw_out Temp'] = Hot_intemp1
   ws['Cold In Temp'] = Temp_in
   ws['Cold Out Temp.'] = Cold_outtemp
   ws['Cold Out Spec.'] = Cold_Spec
   ws['Div. N'] = N
   ws['Fouling Factor'] = Fouling_F 
   ws['La_Heat KJ'] = La_Heat * 4.184
   ws['Heat Capa (KW)'] = ToL_Q / 859.8
   ws['Surf Area (m^2)'] = T_surface 
   ws['LMTD'] = LMTD
   ws['Pressure Loss'] = DPT / 1000
   ws['Cold (Kg/h)'] = Cold_Q
   ws['Gw (Kg/h)'] = Hot_Q
   ws['Cold in V (m/s)'] = Cold_VC 
   ws['Cold Out V(m/s)'] = Cold_V
   ws['TC1 열전도도'] = C_TC1 * 1.163
   ws['TC'] = C_TC *1.163
   ws['CP1  비열'] = C_CP1 * 4.184
   ws['CP'] = C_CP  * 4.184
   ws['Heat Trasnfer kcal/m2.hr'] = ToTal_K																																		
   ws['SG1 (비중)'] = SG1         # Gas in
   ws['SG'] = 1/C_VO/1000	# Gas	out																					
   ws['Spec Capa (KW)'] = ToL_Spec																								
   ws['VI1(Kinetic Viscosity)'] = C_VI1 * 10000
   ws['VI(Kinetic Viscosity)'] = C_VI	* 10000																						
   ws['Volum rate1 (m3/sec)'] = Volume_Rate1 / 3600
   ws['Volum rate (m3/sec)'] = Volume_Rate	/ 3600
   ws['project'] = A   # Project
   ws['Date'] = datetime.today().strftime("%Y. %m .%d")  # YYYY.mm.dd 형태의 출력																						
   ws['Custom'] = B	# Custom	
   ws['DpgT'] = DpgT/1000
   ws['GW_V'] = Hot_VC
   ws['ToL_Spec'] = ToL_Spec
   ws['Plate Pitch'] = t_pitch
   ws['-'] = 100-surface_pe
   ws['L'] = L # Pass 
   # ws['D13'] = GW_cP
   ws['Type Shell&Tube, Plate&shell'] = C   # Type Shell&Tube, Plate&shell
   ws['LNG'] = D   # LNG
   ws['GW'] = E   # GW 
   ws['H_CP1'] = H_CP1 * 4.184 #
   ws['H_CP'] = H_CP * 4.184 # 
   ws['H_TC1'] = H_TC1   #
   ws['H_TC'] = H_TC   #
   ws['H_VO1'] = H_VO1   #
   ws['H_VO'] = H_VO   #
   ws['Cold_VC'] = Cold_VC   #
   ws['Cold_VCT'] = Cold_VCT   #
   ws['Hot_VC'] = Hot_VC   #
   ws['Hot_VCT'] = Hot_VCT   #
   ws['C_cPC'] = C_cPC   #
   ws['H_cPC'] = H_cPC   #
   ws['H_cPT'] = H_cPT   #
   ws['TT2'] = TT2  #

   print(res)
   return jsonify({"image": f'data:image/png;base64,{img_base64}', "printData" : res, "excelData" : ws})

@app.route("/api/plot")
def plot():
    # 예시 그래프 데이터
    x = [1, 2, 3, 4]
    y = [10, 20, 25, 30]

    # 이미지 생성
    fig, ax = plt.subplots()
    ax.plot(x, y)
    ax.set_title("Sample Plot")
    ax.set_xlabel("X-axis")
    ax.set_ylabel("Y-axis")

    # 버퍼에 PNG 저장
    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    buf.seek(0)
    plt.close(fig)

    # Base64 인코딩
    img_base64 = base64.b64encode(buf.read()).decode("utf-8")
    return jsonify({"image": img_base64})

def start_flask():
    app.run(host="127.0.0.1", port=5000)

if __name__ == "__main__":
    t = threading.Thread(target=start_flask, daemon=True)
    t.start()
    webview.create_window("React + Python", "http://localhost:5173")
    webview.start()